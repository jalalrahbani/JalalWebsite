import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import tkinter.font as tkFont
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
from scipy.spatial.distance import euclidean
from scipy.stats import ttest_ind
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import seaborn as sns

# Function to calculate speed, directionality, and other metrics
def analyze_cell_track(df, pixel_size):
    results = {}
    df = df.sort_values(by='slice number')  # Ensure data is sorted by slice number

    df['dx'] = df['X coordinate'].diff().fillna(0) * pixel_size
    df['dy'] = df['Y coordinate'].diff().fillna(0) * pixel_size
    df['distance'] = np.sqrt(df['dx']**2 + df['dy']**2).fillna(0)
    df['time_diff'] = df['time (min)'].diff().bfill()  # Calculate time differences using backward fill
    df['instantaneous_speed'] = df.apply(lambda row: row['distance'] / row['time_diff'] if row['time_diff'] != 0 else 0, axis=1)
    df['direction'] = np.arctan2(df['dy'], df['dx']).fillna(0)
    df['turning_angle'] = df['direction'].diff().fillna(0)

    results['total_distance'] = df['distance'].sum()
    results['net_displacement'] = euclidean((df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size), (df['X coordinate'].iloc[-1] * pixel_size, df['Y coordinate'].iloc[-1] * pixel_size))
    results['average_speed'] = df['instantaneous_speed'].mean()
    results['instantaneous_speed'] = df['instantaneous_speed'].tolist()
    results['direction'] = df['direction'].tolist()
    results['turning_angles'] = df['turning_angle'].tolist()

    # Calculate directionality ratio over time
    directionality_time = []
    for i in range(1, len(df)):
        net_disp = euclidean((df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size), (df['X coordinate'].iloc[i] * pixel_size, df['Y coordinate'].iloc[i] * pixel_size))
        total_dist = df['distance'].iloc[:i+1].sum()
        directionality_time.append(net_disp / total_dist if total_dist != 0 else np.nan)
    results['directionality_over_time'] = directionality_time

    # Calculate directionality ratio over distance
    distance_bins = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]  # Bins in the scaled units (e.g., micrometers)
    distance_bins = [bin_value * pixel_size for bin_value in distance_bins]  # Scale the bins
    directionality_distance = []
    total_distance_traveled = 0
    bin_index = 0
    for i in range(1, len(df)):
        total_distance_traveled += df['distance'].iloc[i]
        while bin_index < len(distance_bins) and total_distance_traveled >= distance_bins[bin_index]:
            net_disp = euclidean((df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size), (df['X coordinate'].iloc[i] * pixel_size, df['Y coordinate'].iloc[i] * pixel_size))
            directionality_distance.append(net_disp / total_distance_traveled if total_distance_traveled != 0 else np.nan)
            bin_index += 1
            if bin_index >= len(distance_bins):
                break
    results['directionality_over_distance'] = directionality_distance

    results['directionality'] = results['net_displacement'] / results['total_distance'] if results['total_distance'] != 0 else np.nan
    results['persistence_time'] = (df['turning_angle'].abs() < np.pi/4).sum() * df['time_diff'].mean()
    results['path_tortuosity'] = results['total_distance'] / results['net_displacement'] if results['net_displacement'] != 0 else np.nan

    # Correct MSD calculation
    msd = []
    num_points = len(df)
    for tau in range(1, num_points):
        displacements = [(df['X coordinate'].iloc[i+tau] - df['X coordinate'].iloc[i])**2 + 
                         (df['Y coordinate'].iloc[i+tau] - df['Y coordinate'].iloc[i])**2 
                         for i in range(num_points - tau)]
        msd.append(np.mean(displacements) * pixel_size**2)
    results['mean_squared_displacement'] = msd

    # Debugging: Print MSD calculations
    print(f"Track {df['track number'].iloc[0]} MSD: {results['mean_squared_displacement']}")

    results['data'] = df

    return results

# Function to process all sheets in the Excel file
def process_excel_file(file_path, pixel_size):
    xl = pd.ExcelFile(file_path)
    all_results = {}

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        track_ids = df['track number'].unique()
        sheet_results = {}

        for track_id in track_ids:
            track_df = df[df['track number'] == track_id].copy()
            track_results = analyze_cell_track(track_df, pixel_size)
            sheet_results[track_id] = track_results

        all_results[sheet_name] = sheet_results

    return all_results

# Function to compare conditions
def compare_conditions(results):
    comparison = {}
    conditions = list(results.keys())

    for i, condition1 in enumerate(conditions):
        for j, condition2 in enumerate(conditions):
            if i >= j:
                continue
            comparison[f'{condition1} vs {condition2}'] = {}
            for metric in results[condition1][list(results[condition1].keys())[0]].keys():
                if metric == 'data':
                    continue
                condition1_values = [results[condition1][track_id][metric] for track_id in results[condition1].keys()]
                condition2_values = [results[condition2][track_id][metric] for track_id in results[condition2].keys()]
                if isinstance(condition1_values[0], list) and isinstance(condition2_values[0], list):
                    flattened_condition1_values = [item for sublist in condition1_values for item in sublist]
                    flattened_condition2_values = [item for sublist in condition2_values for item in sublist]
                    if len(flattened_condition1_values) > 0 and len(flattened_condition2_values) > 0:
                        t_stat, p_value = ttest_ind(flattened_condition1_values, flattened_condition2_values, equal_var=False)
                        comparison[f'{condition1} vs {condition2}'][metric] = {'t_stat': t_stat, 'p_value': p_value}
                    else:
                        comparison[f'{condition1} vs {condition2}'][metric] = {'t_stat': np.nan, 'p_value': np.nan}
                else:
                    t_stat, p_value = ttest_ind(condition1_values, condition2_values, equal_var=False)
                    comparison[f'{condition1} vs {condition2}'][metric] = {'t_stat': t_stat, 'p_value': p_value}

    return comparison

# Function to plot various metrics for each condition and save the plots
def plot_metrics(results, output_dir):
    metrics_to_plot = [
        'total_distance', 'net_displacement', 'average_speed', 'path_tortuosity',
        'directionality_over_time', 'directionality_over_distance', 'mean_squared_displacement'
    ]

    for metric in metrics_to_plot:
        plt.figure(figsize=(12, 8))

        if metric in ['directionality_over_time', 'directionality_over_distance', 'mean_squared_displacement']:
            condition_colors = {condition: color for condition, color in zip(results.keys(), sns.color_palette("husl", len(results)))}
            for condition, tracks in results.items():
                for track_id, track_metrics in tracks.items():
                    plt.plot(track_metrics[metric], color=condition_colors[condition], alpha=0.5)
            for condition, color in condition_colors.items():
                plt.plot([], [], color=color, label=condition)
            plt.legend(title='Conditions')
        else:
            for condition, tracks in results.items():
                values = [tracks[track_id][metric] for track_id in tracks.keys()]
                sns.violinplot(data=values, inner='point')
                sns.swarmplot(data=values, color=".25")

        plt.title(metric.replace('_', ' ').capitalize())
        plt.xlabel('Condition')
        plt.ylabel(metric.replace('_', ' ').capitalize())
        plt.legend()
        plt.tight_layout()

        # Save the plot
        plot_filename = os.path.join(output_dir, f'{metric}.png')
        plt.savefig(plot_filename)
        plt.close()

# Function to save results with alternating colors
def save_results_with_colors(results, comparison, output_file_path):
    global headers
    wb = Workbook()

    # Create the summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    summary_data = []
    for condition, tracks in results.items():
        for track_id, track_metrics in tracks.items():
            summary_data.append({
                'Condition': condition,
                'Track ID': track_id,
                'Total Distance': round(track_metrics['total_distance'], 2),
                'Net Displacement': round(track_metrics['net_displacement'], 2),
                'Average Speed': round(track_metrics['average_speed'], 2),
                'Path Tortuosity': round(track_metrics['path_tortuosity'], 2) if track_metrics['path_tortuosity'] is not None else None,
                'Persistence Time': round(track_metrics['persistence_time'], 2) if track_metrics['persistence_time'] is not None else None
            })

    # Convert summary to DataFrame and add to worksheet
    df_summary = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)

    # Loop through each condition and its metrics
    for condition, metrics in results.items():
        sheet_name = f'{condition}_Metrics'
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]  # Excel sheet names cannot exceed 31 characters
        ws = wb.create_sheet(sheet_name)

        # Define headers
        headers = [
            'Track ID', 'Slice Number', 'X Coordinate', 'Y Coordinate', 'Time (min)', 'dx (um)', 'dy (um)', 
    'Distance (um)', 'Instantaneous Speed (um/min)', 'Direction (rad)', 'Turning Angle (rad)', 
    'Directionality Over Time', 'Directionality Over Distance', 'Path Tortuosity', 
    'Persistence Time (min)', 'Total Distance (um)', 'Net Displacement (um)', 'Average Speed (um/min)', 
    'Mean Squared Displacement (um^2)'
        ]
        ws.append(headers)

        # Define alternating colors for rows
        fill_grey = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Loop through track data and write each row with alternating colors
        for track_id, track_metrics in metrics.items():
            fill = fill_grey if track_id % 2 == 0 else fill_white
            df = track_metrics['data']
            df['Track ID'] = track_id

            # Handle the directionality metrics and make sure lengths match with the dataframe
            directionality_over_time = track_metrics['directionality_over_time']
            if len(directionality_over_time) < len(df):
                directionality_over_time.extend([np.nan] * (len(df) - len(directionality_over_time)))
            df['Directionality Over Time'] = directionality_over_time

            directionality_over_distance = track_metrics['directionality_over_distance']
            if len(directionality_over_distance) < len(df):
                directionality_over_distance.extend([np.nan] * (len(df) - len(directionality_over_distance)))
            df['Directionality Over Distance'] = directionality_over_distance

            # Add calculated metrics to the DataFrame
            df['Path Tortuosity'] = track_metrics['path_tortuosity']
            df['Persistence Time'] = track_metrics['persistence_time']
            df['Total Distance'] = track_metrics['total_distance']
            df['Net Displacement'] = track_metrics['net_displacement']
            df['Average Speed'] = track_metrics['average_speed']

            # Ensure length match for Mean Squared Displacement (MSD)
            mean_squared_displacement = track_metrics['mean_squared_displacement']
            if len(mean_squared_displacement) < len(df):
                mean_squared_displacement.extend([np.nan] * (len(df) - len(mean_squared_displacement)))
            df['Mean Squared Displacement'] = mean_squared_displacement

            # Select relevant columns and apply rounding to float columns
            df = df[[
                'Track ID', 'slice number', 'X coordinate', 'Y coordinate', 'time (min)', 'dx', 'dy', 'distance',
                'instantaneous_speed', 'direction', 'turning_angle', 'Directionality Over Time', 'Directionality Over Distance',
                'Path Tortuosity', 'Persistence Time', 'Total Distance', 'Net Displacement', 'Average Speed',
                'Mean Squared Displacement'
            ]].round(2)  # Round all numeric values to 2 decimal places

            # Add rows to the Excel worksheet
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
                for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    for c in cell:
                        c.fill = fill  # Apply alternating row colors

    # Create a comparison sheet
    for comparison_key, comparison_metrics in comparison.items():
        sheet_name = f'{comparison_key}_comparison'
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]  # Handle Excel sheet name limitation
        ws = wb.create_sheet(sheet_name)
        df_comparison = pd.DataFrame.from_dict(comparison_metrics, orient='index')

        # Write the comparison data
        for r in dataframe_to_rows(df_comparison.reset_index(), index=False, header=True):
            ws.append(r)

    # Save the workbook to the specified output file
    wb.save(output_file_path)

    
#GUI for Tkinter App
class CellMigrationAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cell Migration Analyzer")
        self.root.geometry("800x600")

        # Create a custome style
        style = ttk.Style()

        # Configure the Notebook (tabs) to have larger size and round edges
        style.theme_use('clam')  # Use the 'clam' theme as a base
        style.configure("TNotebook", tabmargins=[2, 5, 2, 0])  # Add margins around the tabs
        style.configure("TNotebook.Tab", padding=[20, 10], font=('Helvetica', 12))  # Make tabs larger, more padding
        style.map("TNotebook.Tab", background=[("selected", "#87CEEB")], foreground=[("selected", "black")])  # Custom color when selected

        # Style buttons with rounded corners
        style.configure("TButton", font=('Helvetica', 10), relief="flat", background="#ADD8E6", borderwidth=1)
        style.map("TButton", relief=[("active", "flat")], background=[("active", "#87CEEB")])

        # File path and data storage
        self.file_path = None
        self.dataframes = {}
        self.generated_plots = []  # To store generated plots for visualization
        self.results = None  # To store analysis results
        self.current_plot_index = 0  # Track the current plot being shown
        self.is_data_loaded = False  # Track whether data has been loaded

        # Add notebook (tabs)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(pady=10, expand=True)

        # Analyze method should be defined before calling `setup_tabs`
        self.setup_tabs()

    def analyze(self):
        """Run the analysis and generate plots."""
        pixel_size = self.pixel_size_entry.get()
        if self.file_path and pixel_size:
            try:
                pixel_size = float(pixel_size)
                self.results = process_excel_file(self.file_path, pixel_size)
                comparison = compare_conditions(self.results)

                # Generate the plots and store them for visualization
                self.generated_plots = self.plot_metrics(self.results)

                # Show the first plot immediately
                self.current_plot_index = 0
                self.show_plot(self.current_plot_index)

                messagebox.showinfo("Analysis Complete", "Analysis completed successfully.")
            except ValueError:
                messagebox.showerror("Input Error", "Please enter a valid pixel size.")
        else:
            messagebox.showerror("Missing Information", "Please select a file and enter pixel size.")

    def setup_tabs(self):
        # Tab 1: Image Loading and Preprocessing
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text="Image Preprocessing-COMING SOON")

        # Add labels, buttons, or widgets specific to image loading and preprocessing in Tab 1 here.
        # Placeholder Label
        tk.Label(self.tab1, text="Load Imaging File and Perform Preprocessing").pack(pady=10)

        # Tab 2: AI Automated Tracking
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text="AI Automated Tracking-COMING SOON")

        # Placeholder Label
        tk.Label(self.tab2, text="Load Processed Image and Train AI for Cell Tracking").pack(pady=10)

        # Tab 3: Excel sheet File Selection and Pixel Size
        self.tab3 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3, text="Select Tracked files")

        # Use ttk.Button for modern style and rounder corners
        self.file_button = ttk.Button(self.tab3, text="Select Excel File", command=self.load_file)
        self.file_button.pack(pady=20)

        # Labels can remain tk.Label since they don't support ttk styling
        self.pixel_size_label = tk.Label(self.tab3, text="Enter Pixel Size (micrometers):", font=("Helvetica", 10))
        self.pixel_size_label.pack(pady=5)

        # Entry fields can be ttk.Entry for a consistent look
        self.pixel_size_entry = ttk.Entry(self.tab3, font=("Helvetica", 10))
        self.pixel_size_entry.pack(pady=10)

        # Use ttk.Button for Analyze button
        self.analyze_button = ttk.Button(self.tab3, text="Analyze", command=self.analyze)
        self.analyze_button.pack(pady=20)

        # Tab 4: Inspect and Edit Data
        self.tab4 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab4, text="Inspect Data")

        self.sheet_listbox = tk.Listbox(self.tab4)  # Define the listbox
        self.sheet_listbox.pack(pady=10, fill=tk.BOTH, expand=True)  # Pack it into the tab

        # Dropdown to switch between Original Data and Calculated Results
        self.data_view_choice = tk.StringVar(value="Original Data")
        self.data_view_menu = ttk.OptionMenu(self.tab4, self.data_view_choice, "Original Data", "Original Data", "Calculated Results", command=self.view_sheet)
        self.data_view_menu.pack(pady=10)

        # Create a frame to hold the treeview and scrollbars
        self.tree_frame = tk.Frame(self.tab4)
        self.tree_frame.pack(pady=10, fill=tk.BOTH, expand=True)

        # Create the Treeview widget inside the frame
        self.tree = ttk.Treeview(self.tree_frame)
        self.tree.grid(row=0, column=0, sticky='nsew')  # Stretch to fit the frame

        # Scrollbars
        self.tree_scroll_y = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree_scroll_y.grid(row=0, column=1, sticky='ns')  # Stick to the right of the tree

        self.tree_scroll_x = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree_scroll_x.grid(row=1, column=0, sticky='ew')  # Stick to the bottom of the tree

        # Configure treeview to work with scrollbars
        self.tree.configure(yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set)

        # Configure the frame to expand properly
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

        # Use ttk.Button for Edit Cell button
        self.edit_button = ttk.Button(self.tab4, text="Edit Selected Cell", command=self.edit_cell)
        self.edit_button.pack(pady=10)  # Pack the button below the treeview

        # Tab 5: Visualization
        self.tab5 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab5, text="Results & Visuals")

        self.plot_title_label = tk.Label(self.tab5, text="", font=('Helvetica', 14))
        self.plot_title_label.pack(pady=10)

        self.canvas_frame = tk.Frame(self.tab5)
        self.canvas_frame.pack(pady=10, fill=tk.BOTH, expand=True)

        # Use ttk.Button for Up and Down buttons
        self.up_button = ttk.Button(self.tab5, text="BACK", command=self.show_previous_plot)
        self.up_button.pack(side=tk.LEFT, padx=5)

        self.down_button = ttk.Button(self.tab5, text="NEXT", command=self.show_next_plot)
        self.down_button.pack(side=tk.RIGHT, padx=5)

        # Add Save Plot and Save Results buttons using ttk.Button
        self.save_results_label = tk.Label(self.tab5, text="Click the button below to save the current Plot as .jpg:")
        self.save_results_label.pack(pady=2)  # Add some padding to space it out
        self.save_button = ttk.Button(self.tab5, text="Save Plot", command=self.save_plot)
        self.save_button.pack(pady=10)

        # Add Save Results button and label
        self.save_results_label = tk.Label(self.tab5, text="Click the button below to save the analysis results in .xlx file:")
        self.save_results_label.pack(pady=2)  # Add some padding to space it out
        self.save_results_button = ttk.Button(self.tab5, text="Export Results", command=self.save_results)
        self.save_results_button.pack(pady=10)

        # Tab 6: CellTrack INFO
        self.tab6 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab6, text="CellTrack INFO")

        # Documentation text area
        self.info_text = tk.Text(self.tab6, wrap=tk.WORD, width=100, height=30)
        self.info_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        # Scrollbars for the info_text widget
        self.info_scroll_y = ttk.Scrollbar(self.tab6, orient="vertical", command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=self.info_scroll_y.set)
        self.info_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        # Insert detailed documentation about the app and its functionality
        self.insert_documentation()

    def insert_documentation(self):
        """Inserts the detailed documentation in the CellTrack INFO tab."""
        
        # Define text tags for formatting
        self.info_text.tag_configure("title", font=("Helvetica", 16, "bold"), justify="center")
        self.info_text.tag_configure("header", font=("Helvetica", 14, "bold"))
        self.info_text.tag_configure("subheader", font=("Helvetica", 12, "italic"))
        self.info_text.tag_configure("normal", font=("Helvetica", 10), wrap="word")
        self.info_text.tag_configure("formula", font=("Helvetica", 10, "italic"))
        self.info_text.tag_configure("space", font=("Helvetica", 10), lmargin1=20, lmargin2=20)
        
        # Insert the formatted text into the Text widget
        self.info_text.insert(tk.END, "\n --CellTrack Documentation-- \n", "title")

        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        self.info_text.insert(tk.END, "\n Introduction: \n", "header")
        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")

        self.info_text.insert(tk.END, "\nThis application analyzes cell migration from Excel files that track cell coordinates over time. It processes data from each track, calculates various metrics such as speed, directionality, path tortuosity, and more, and provides visualizations and statistical comparisons between conditions\n", "normal")
        
        self.info_text.insert(tk.END, "\nUsing the App:\n", "header")

        self.info_text.insert(tk.END, "1. Select the Excel File\n", "subheader")
        self.info_text.insert(tk.END, "Click on 'Select File' and choose your Excel file containing the data. After the data is loaded a popup window will confirm fact and the data is now visible in the Inspect tab\n", "normal")

        self.info_text.insert(tk.END, "\n2. Enter Pixel Size\n", "subheader")
        self.info_text.insert(tk.END, "To then analyze the data, enter the pixel size in micrometers in the same tab and click 'Analyze' to start processing the data.\n", "normal")

        self.info_text.insert(tk.END, "\n3. Inspect the data\n", "subheader")
        self.info_text.insert(tk.END, "In the 'Inspect Data' tab, view the original or calculated results from the dropdown menu. You can edit individual cell data by highlighting/Selecting a cell and clicking 'Edit Selected Cell.\n", "normal")

        self.info_text.insert(tk.END, "\n4. Export Results and Graphs\n", "subheader")
        self.info_text.insert(tk.END, "In the 'Results & Visuals' tab, browse through visualizations and save plots.\n", "normal")

        self.info_text.insert(tk.END, "\nExcel Sheet Requirements\n", "header")
        self.info_text.insert(tk.END, "The input Excel sheet must adhere to the following structure for the app to work correctly: \n - Each sheet should represent a condition or experiment. \n- Columns required: 'track number', 'slice number', 'X coordinate', 'Y coordinate', 'time (min)'. \n- Each row represents a point in time for a particular track (e.g., slice number 1, 2, 3). \n- Time must be recorded in minutes. \n- X and Y coordinates should represent pixel values, which will be converted into real-world units using the provided pixel size.\n", "normal")

        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        self.info_text.insert(tk.END, "\n Metrics Calculated: \n", "header")
        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")

        # Total Distance
        self.info_text.insert(tk.END, "\n- Total Distance (μm): ", "subheader")
        self.info_text.insert(tk.END, "The cumulative distance traveled along the entire path, measured in micrometers (μm).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Total Distance = Σ √((X(i) - X(i-1))² + (Y(i) - Y(i-1))²), in μm\n", "formula")

        # Net Displacement
        self.info_text.insert(tk.END, "\n- Net Displacement (μm): ", "subheader")
        self.info_text.insert(tk.END, "The straight-line distance between the start and end of a track, measured in micrometers (μm).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Net Displacement = √((X_end - X_start)² + (Y_end - Y_start)²), in μm\n", "formula")

        # Average Speed
        self.info_text.insert(tk.END, "\n- Average Speed (μm/min): ", "subheader")
        self.info_text.insert(tk.END, "The average speed of the cell over the entire track, calculated as the total "
                                    "distance divided by the total time, measured in micrometers per minute (μm/min).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Average Speed = Total Distance / Total Time, in μm/min\n", "formula")

        # Path Tortuosity
        self.info_text.insert(tk.END, "\n- Path Tortuosity: ", "subheader")
        self.info_text.insert(tk.END, "Path tortuosity measures how much the path deviates from a straight line, "
                                    "calculated as the ratio of Total Distance to Net Displacement (dimensionless).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Tortuosity = Total Distance / Net Displacement\n", "formula")

        # Mean Squared Displacement (MSD)
        self.info_text.insert(tk.END, "\n- Mean Squared Displacement (MSD, μm²): ", "subheader")
        self.info_text.insert(tk.END, "MSD is a measure of how far a cell moves from its origin, averaged over different "
                                    "time intervals, measured in square micrometers (μm²).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "MSD(τ) = ⟨(X(t + τ) - X(t))² + (Y(t + τ) - Y(t))²⟩, in μm²\n", "formula")

        # Directionality Over Time
        self.info_text.insert(tk.END, "\n- Directionality Over Time: ", "subheader")
        self.info_text.insert(tk.END, "This metric tracks how the cell’s directional persistence changes over time. It "
                                    "is calculated by examining how the ratio of Net Displacement to Total Distance evolves over time.\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Directionality(t) = Net Displacement(t) / Total Distance(t)\n", "formula")

        # Directionality Over Distance
        self.info_text.insert(tk.END, "\n- Directionality Over Distance: ", "subheader")
        self.info_text.insert(tk.END, "Similar to directionality over time, this metric assesses how persistent the cell "
                                    "is in maintaining directionality as a function of the total distance traveled.\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Directionality(d) = Net Displacement(d) / Total Distance(d)\n", "formula")

        # Persistence Time
        self.info_text.insert(tk.END, "\n- Persistence Time (min): ", "subheader")
        self.info_text.insert(tk.END, "Persistence time is the total time during which the cell maintains relatively "
                                    "straight movement, calculated by summing up the time intervals where the turning "
                                    "angle is below a specific threshold, measured in minutes.\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Persistence Time = Σ(time where |Turning Angle| < threshold), in minutes\n", "formula")

        # Instantaneous Speed
        self.info_text.insert(tk.END, "\n- Instantaneous Speed (μm/min): ", "subheader")
        self.info_text.insert(tk.END, "The speed calculated between consecutive time points, measured in micrometers per minute (μm/min).\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Instantaneous Speed = Distance / Time, in μm/min\n", "formula")

        # Turning Angles
        self.info_text.insert(tk.END, "\n- Turning Angles (radians): ", "subheader")
        self.info_text.insert(tk.END, "The change in direction between consecutive time points, helping to assess how "
                                    "much the cell changes its movement direction, measured in radians.\n", "normal")
        self.info_text.insert(tk.END, "\nFormula:\n", "normal")
        self.info_text.insert(tk.END, "Turning Angle = arctan2(ΔY, ΔX), in radians\n", "formula")


        # Wrapping up
        self.info_text.insert(tk.END, "\nThis concludes the explanation of metrics calculated in this application. "
                                      "These metrics provide comprehensive insight into cell migration behavior and enable "
                                      "statistical comparisons between different conditions.\n", "normal")
        
        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        # Explanation about the exported Excel sheet
        self.info_text.insert(tk.END, " Results : Exported Excel Sheet Overview: \n", "header")
        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        
        self.info_text.insert(tk.END, "\nOnce the analysis is complete, you can export the results into an Excel file.\n", "normal")
        
        self.info_text.insert(tk.END, "\n1. Summary Sheet:\n", "subheader")
        self.info_text.insert(tk.END, "The first sheet in the exported Excel file is a 'Summary' sheet that contains a concise overview of the most important metrics for each track across all conditions.\n", "normal")
        self.info_text.insert(tk.END, "It includes the following columns:\n", "normal")
        self.info_text.insert(tk.END, "- Condition: The experimental condition associated with each track.\n", "normal")
        self.info_text.insert(tk.END, "- Track ID: The unique identifier for each cell track.\n", "normal")
        self.info_text.insert(tk.END, "- Total Distance: The total distance covered by each track.\n", "normal")
        self.info_text.insert(tk.END, "- Net Displacement: The straight-line distance between the start and end of the track.\n", "normal")
        self.info_text.insert(tk.END, "- Average Speed: The average speed across the track.\n", "normal")
        self.info_text.insert(tk.END, "- Path Tortuosity: A ratio indicating how direct the movement is.\n", "normal")
        self.info_text.insert(tk.END, "- Persistence Time: The time during which the cell moved in a relatively straight path.\n", "normal")

        self.info_text.insert(tk.END, "\n2. Condition-Specific Sheets:\n", "subheader")
        self.info_text.insert(tk.END, "Each condition will also have its own sheet, named according to the condition. If the name exceeds 31 characters (Excel's sheet name limit), it will be truncated.\n", "normal")
        self.info_text.insert(tk.END, "These sheets contain detailed data for every track within that condition. The columns in these sheets are as follows:\n", "normal")
        self.info_text.insert(tk.END, "- Track ID: The unique identifier for each track.\n", "normal")
        self.info_text.insert(tk.END, "- Slice Number: The time point or frame number for each track's measurement.\n", "normal")
        self.info_text.insert(tk.END, "- X Coordinate & Y Coordinate: The cell's position at each time point.\n", "normal")
        self.info_text.insert(tk.END, "- Time (min): The time in minutes associated with each measurement.\n", "normal")
        self.info_text.insert(tk.END, "- dx, dy: The change in X and Y coordinates between consecutive time points.\n", "normal")
        self.info_text.insert(tk.END, "- Distance: The distance traveled between consecutive points.\n", "normal")
        self.info_text.insert(tk.END, "- Instantaneous Speed: The speed between consecutive points.\n", "normal")
        self.info_text.insert(tk.END, "- Direction: The angle of movement between consecutive points.\n", "normal")
        self.info_text.insert(tk.END, "- Turning Angle: The change in direction between consecutive points.\n", "normal")
        self.info_text.insert(tk.END, "- Directionality Over Time: The ratio of Net Displacement to Total Distance at different time points.\n", "normal")
        self.info_text.insert(tk.END, "- Directionality Over Distance: Similar to the above, but calculated based on distance instead of time.\n", "normal")
        self.info_text.insert(tk.END, "- Path Tortuosity: The ratio of Total Distance to Net Displacement for each track.\n", "normal")
        self.info_text.insert(tk.END, "- Persistence Time: The time during which the cell is persistently moving in a single direction.\n", "normal")
        self.info_text.insert(tk.END, "- Mean Squared Displacement (MSD): The average squared distance from the origin, calculated over multiple time intervals.\n", "normal")

        self.info_text.insert(tk.END, "\nThese condition-specific sheets provide more detailed information than the summary sheet, making it easy to dive deeper into the behavior of individual cells.\n", "normal")
        self.info_text.insert(tk.END, "\n\n", "normal")

        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        self.info_text.insert(tk.END, "\n Statistical Analysis \n", "header")
        # Add a line break for clarity
        self.info_text.insert(tk.END, "\n\n---------------------------------------------------------------------------------------------------------------------------------------------------------------\n\n", "normal")
        
        self.info_text.insert(tk.END, "\n- T-Test Comparisons: ", "subheader")
        self.info_text.insert(tk.END, 
            "The application performs a two-sample t-test (Welch’s t-test) to statistically compare the metrics between conditions. "
            "This test is applied to determine whether there is a significant difference between the selected conditions for each calculated metric.\n"
            , "normal")
            
        self.info_text.insert(tk.END, 
            "\n- How It Works: ", "subheader")
        self.info_text.insert(tk.END, 
            "The t-test is calculated for each metric (e.g., Total Distance, Net Displacement, Average Speed) "
            "between pairs of conditions provided in the dataset. If a metric contains multiple values (such as a time series like Directionality Over Time), "
            "the values are flattened to form a single set for each condition before applying the statistical test.\n"
            , "normal")

        self.info_text.insert(tk.END, 
            "\n- Where the Results Are Saved: ", "subheader")
        self.info_text.insert(tk.END, 
            "The statistical comparison results are saved in a separate sheet in the Excel export file. Each sheet corresponding to a comparison is named in the format 'Condition1 vs Condition2 Comparison'. "
            "In these sheets, the t-statistic and p-value for each metric are displayed. The p-value indicates whether there is a statistically significant difference between the two conditions for that metric.\n"
            , "normal")

        self.info_text.insert(tk.END, 
            "\n- How to Interpret the Results: ", "subheader")
        self.info_text.insert(tk.END, 
            "In the comparison sheets, a low p-value (typically < 0.05) suggests that there is a significant difference between the two conditions for that metric. "
            "The t-statistic gives an indication of the magnitude and direction of the difference. A positive t-statistic suggests that the metric is higher for the first condition compared to the second, while a negative t-statistic suggests the opposite.\n"
            , "normal")

        self.info_text.configure(state="disabled")  # Disable editing of the text once it's inserted
  
    def update_tab_color(self, is_data_loaded):
        """Change the tab label depending on whether data is loaded."""
        if is_data_loaded:
            self.notebook.tab(1, text="Inspect Data (Loaded)")
        else:
            self.notebook.tab(1, text="Inspect Data (Empty)")

    # Only update tab color when loading the file, not when switching views
    def load_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.file_path:
            xl = pd.ExcelFile(self.file_path)
            self.dataframes = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}
            self.sheet_listbox.delete(0, tk.END)
            for sheet in xl.sheet_names:
                self.sheet_listbox.insert(tk.END, sheet)
            messagebox.showinfo("File Selected", f"Loaded {self.file_path}")

            # Only update tab title when loading a new file
            self.is_data_loaded = True
            self.update_tab_color(self.is_data_loaded)

            # Ensure window size remains fixed
            self.adjust_window_size()

    def show_previous_plot(self):
        """Show the previous plot in the list."""
        if self.generated_plots and self.current_plot_index > 0:
            self.current_plot_index -= 1
            self.show_plot(self.current_plot_index)

        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def show_next_plot(self):
        """Show the next plot in the list."""
        if self.generated_plots and self.current_plot_index < len(self.generated_plots) - 1:
            self.current_plot_index += 1
            self.show_plot(self.current_plot_index)

        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def show_plot(self, plot_index):
        """Display a specific plot from the list of generated plots."""
        if self.generated_plots and 0 <= plot_index < len(self.generated_plots):
            plot, title = self.generated_plots[plot_index]

            # Clear previous plot
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()

            self.plot_title_label.config(text=title.replace('_', ' ').capitalize())

            canvas = FigureCanvasTkAgg(plot, master=self.canvas_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.adjust_window_size()  # Adjust the window size dynamically after loading the file
    
    # Prevent tab title from changing when switching data view
    def view_sheet(self, choice=None):
        selected_sheet = self.sheet_listbox.get(tk.ACTIVE)  # Get the selected sheet from the listbox
        if selected_sheet:
            # Reset the dropdown to have both options available
            self.data_view_menu['menu'].delete(0, 'end')
            self.data_view_menu['menu'].add_command(label="Original Data", command=lambda: self.view_sheet("Original Data"))
            self.data_view_menu['menu'].add_command(label="Calculated Results", command=lambda: self.view_sheet("Calculated Results"))
            self.data_view_choice.set(choice if choice else "Original Data")

            # Refresh the dropdown display
            self.notebook.select(self.tab4)

            # Original Data view
            if self.data_view_choice.get() == "Original Data":
                if selected_sheet in self.dataframes:
                    df = self.dataframes[selected_sheet]
                    self.selected_df = df  # Store the DataFrame as self.selected_df for editing
                    self.display_dataframe_in_treeview(df)
                else:
                    messagebox.showerror("Data Error", f"No original data found for {selected_sheet}.")

            # Calculated Results view
            elif self.data_view_choice.get() == "Calculated Results" and self.results:
                sheet_results = self.results.get(selected_sheet, {})
                if sheet_results:
                    combined_data = []
                    for track_id, track_metrics in sheet_results.items():
                        msd = ', '.join([f'{x:.2f}' for x in track_metrics['mean_squared_displacement']])
                        directionality_over_time = ', '.join([f'{x:.2f}' for x in track_metrics['directionality_over_time']])
                        directionality_over_distance = ', '.join([f'{x:.2f}' for x in track_metrics['directionality_over_distance']])

                        combined_data.append({
                            'Track ID': track_id,
                            'Total Distance': round(track_metrics['total_distance'], 2),
                            'Net Displacement': round(track_metrics['net_displacement'], 2),
                            'Average Speed': round(track_metrics['average_speed'], 2),
                            'Path Tortuosity': round(track_metrics['path_tortuosity'], 2) if track_metrics['path_tortuosity'] else None,
                            'Mean Squared Displacement': msd,
                            'Directionality Over Time': directionality_over_time,
                            'Directionality Over Distance': directionality_over_distance,
                            'Persistence Time': round(track_metrics['persistence_time'], 2) if track_metrics['persistence_time'] else None
                        })

                    combined_df = pd.DataFrame(combined_data)
                    self.selected_df = combined_df  # Store the DataFrame as self.selected_df for editing
                    self.display_dataframe_in_treeview(combined_df)
                else:
                    messagebox.showerror("No Results", "No calculated results for this sheet.")
            else:
                messagebox.showerror("No Selection", "Please select a valid sheet or result type.")

    def combine_results(self):
        """Combine the analysis results into a DataFrame for display."""
        combined_data = []
        for sheet_name, sheet_results in self.results.items():
            for track_id, track_metrics in sheet_results.items():
                combined_data.append({
                    'Sheet Name': sheet_name,
                    'Track ID': track_id,
                    'Total Distance': track_metrics['total_distance'],
                    'Net Displacement': track_metrics['net_displacement'],
                    'Average Speed': track_metrics['average_speed'],
                    'Path Tortuosity': track_metrics['path_tortuosity']
                    # Add any other metrics you wish to display here
                })
        return pd.DataFrame(combined_data)
    
    def display_dataframe_in_treeview(self, df):
        """Display the DataFrame in the TreeView widget with scrollbars and fixed size."""
        # Clear existing treeview contents
        self.tree.delete(*self.tree.get_children())

        # Round all float values to 2 decimal places in the DataFrame before displaying
        df = df.round(2)  # Apply rounding here

        # Set the new columns and headers in the treeview
        self.tree["columns"] = list(df.columns)  # Set columns to match the DataFrame's columns
        self.tree["show"] = "headings"  # Show only the headings, not the default empty column

        # Adjust column width dynamically based on the content
        for col in df.columns:
            max_width = max(df[col].astype(str).map(len).max(), len(col))  # Adjust based on both column name and content length
            col_width = max(100, min(300, max_width * 10))  # Set a reasonable range for column width (min 100, max 300)

            # Configure the column headings and column widths
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_width, anchor='center')  # Center text in columns

        # Insert the DataFrame rows into the treeview
        for index, row in df.iterrows():
            self.tree.insert("", "end", values=list(row))

        # FIXED SIZE - Keep treeview size constant with scrollbars
        self.tree_frame.config(height=400, width=600)  # Fixed size for the treeview frame

        # Ensure the grid geometry manager is used consistently for the treeview and scrollbars
        self.tree_scroll_y.grid(row=0, column=1, sticky='ns')
        self.tree_scroll_x.grid(row=1, column=0, sticky='ew')

        # Reconfigure scrollbars for the treeview
        self.tree.configure(yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set)
        self.tree_scroll_y.config(command=self.tree.yview)
        self.tree_scroll_x.config(command=self.tree.xview)

    # Function to edit the selected cell
    def edit_cell(self):
        """Allow user to edit the value of the selected cell."""
        selected_item = self.tree.focus()
        if selected_item:
            values = self.tree.item(selected_item, 'values')
            column_names = self.tree["columns"]
            selected_column = tk.simpledialog.askstring("Select Column", f"Which column would you like to edit?\n\n{', '.join(column_names)}")
            
            if selected_column in column_names:
                col_index = column_names.index(selected_column)
                new_value = tk.simpledialog.askstring("Edit Cell", f"Current Value: {values[col_index]}")
                
                if new_value:
                    row_id = self.tree.index(selected_item)
                    col_name = self.tree["columns"][col_index]
                    
                    # Convert new_value to the appropriate type (float or int) if needed
                    try:
                        if self.selected_df[col_name].dtype in [np.float64, np.float32, np.int64, np.int32]:
                            new_value = float(new_value)
                        else:
                            new_value = str(new_value)
                    except ValueError:
                        messagebox.showerror("Invalid Input", "Please enter a valid number.")
                        return

                    # Update the DataFrame and Treeview with the new value
                    self.selected_df.at[row_id, col_name] = new_value
                    updated_values = list(self.selected_df.iloc[row_id])
                    self.tree.item(selected_item, values=updated_values)
            else:
                messagebox.showerror("Invalid Column", "Please select a valid column to edit.")
        else:
            messagebox.showerror("No Selection", "Please select a row to edit.")

        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def analyze(self):
        """Run the analysis and generate plots."""
        pixel_size = self.pixel_size_entry.get()
        if self.file_path and pixel_size:
            try:
                pixel_size = float(pixel_size)
                self.results = process_excel_file(self.file_path, pixel_size)
                comparison = compare_conditions(self.results)

                # Generate the plots and store them for visualization
                self.generated_plots = self.plot_metrics(self.results)

                # Show the first plot immediately
                self.current_plot_index = 0
                self.show_plot(self.current_plot_index)

                messagebox.showinfo("Analysis Complete", "Analysis completed successfully.")
            except ValueError:
                messagebox.showerror("Input Error", "Please enter a valid pixel size.")
        else:
            messagebox.showerror("Missing Information", "Please select a file and enter pixel size.")

    def plot_metrics(self, results):
        """Generate the plots and return them in a list."""
        metrics_to_plot = [
            'total_distance',
            'average_speed',
            'net_displacement',
            'path_tortuosity'
        ]
        
        # Special handling for list-based metrics (directionality, MSD, etc.)
        list_based_metrics = [
            'directionality_over_time',
            'directionality_over_distance',
            'mean_squared_displacement'
        ]

        plots = []
        
        # Plot the simpler metrics
        for metric in metrics_to_plot:
            figure, ax = plt.subplots(figsize=(5, 4))
            plot_data = []
            for condition, tracks in results.items():
                for track_id, track_metrics in tracks.items():
                    plot_data.append({
                        'Condition': condition,
                        metric: track_metrics[metric]
                    })
            
            plot_data = pd.DataFrame(plot_data)
            sns.violinplot(x='Condition', y=metric, data=plot_data, inner='point', ax=ax)
            sns.stripplot(x='Condition', y=metric, data=plot_data, color=".25", ax=ax)

            ax.set_title(metric.replace('_', ' ').capitalize())
            ax.set_xlabel('Condition')
            ax.set_ylabel(metric.replace('_', ' ').capitalize())
            plots.append((figure, metric))  # Store figure and metric as tuple

        # Special case for list-based metrics (directionality, MSD, etc.)
        for metric in list_based_metrics:
            figure, ax = plt.subplots(figsize=(5, 4))
            condition_colors = {condition: color for condition, color in zip(results.keys(), sns.color_palette("husl", len(results)))}
            
            if metric == 'directionality_over_time':
                # X-axis will now represent the actual time points in minutes
                for condition, tracks in results.items():
                    for track_id, track_metrics in tracks.items():
                        time_points = list(range(0, len(track_metrics['directionality_over_time']) * 5, 5))  # Assuming 5 min intervals
                        ax.plot(time_points, track_metrics['directionality_over_time'], color=condition_colors[condition], alpha=0.5)
                ax.set_xlabel('Time (min)')  # Change X-axis label to time in minutes
            
            elif metric == 'directionality_over_distance':
                # X-axis will now represent the distance bins (5, 10, 15, etc.)
                distance_bins = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]  # µm
                for condition, tracks in results.items():
                    for track_id, track_metrics in tracks.items():
                        # Use the actual bins as X-axis values
                        ax.plot(distance_bins[:len(track_metrics['directionality_over_distance'])], track_metrics['directionality_over_distance'], color=condition_colors[condition], alpha=0.5)
                ax.set_xlabel('Distance (µm)')  # Change X-axis label to distance in micrometers
            
            elif metric == 'mean_squared_displacement':  # Add the MSD plotting here
                for condition, tracks in results.items():
                    for track_id, track_metrics in tracks.items():
                        ax.plot(track_metrics[metric], alpha=0.5)
                ax.set_xlabel('Frame')
                ax.set_ylabel('Mean Squared Displacement (µm²)')

            ax.set_title(metric.replace('_', ' ').capitalize())
            ax.set_ylabel(metric.replace('_', ' ').capitalize())
            ax.legend(title='Conditions')
            plots.append((figure, metric))  # Store figure and metric as tuple

        return plots

    def show_plot(self, plot_index):
        """Display a specific plot from the list of generated plots."""
        if self.generated_plots and 0 <= plot_index < len(self.generated_plots):
            plot, title = self.generated_plots[plot_index]

            # Clear previous plot
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()

            self.plot_title_label.config(text=title.replace('_', ' ').capitalize())

            canvas = FigureCanvasTkAgg(plot, master=self.canvas_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def show_next_plot(self):
        """Show the next plot in the list."""
        if self.generated_plots and self.current_plot_index < len(self.generated_plots) - 1:
            self.current_plot_index += 1
            self.show_plot(self.current_plot_index)
        
        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def show_previous_plot(self):
        """Show the previous plot in the list."""
        if self.generated_plots and self.current_plot_index > 0:
            self.current_plot_index -= 1
            self.show_plot(self.current_plot_index)

        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def save_plot(self):
        """Saves the currently displayed plot."""
        if self.generated_plots:
            save_dir = filedialog.askdirectory()
            if save_dir:
                plot, title = self.generated_plots[self.current_plot_index]
                plot.savefig(os.path.join(save_dir, f"{title}.png"))
                messagebox.showinfo("Success", f"Plot saved as {title}.png")
        else:
            messagebox.showerror("No Plots", "No plots to save. Run the analysis first.")

    def save_results(self):
        """Saves the analysis results to an Excel file."""
        if self.results:
            output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            if output_file_path:
                comparison = compare_conditions(self.results)
                save_results_with_colors(self.results, comparison, output_file_path)
                messagebox.showinfo("Success", f"Results saved to {output_file_path}")
        else:
            messagebox.showerror("No Results", "No results to save. Run the analysis first.")
        self.adjust_window_size()  # Adjust the window size dynamically after loading the file

    def adjust_window_size(self):
        """Adjust the window size to fit all content dynamically."""
        self.root.update_idletasks()  # Update all pending tasks
        width = self.root.winfo_reqwidth()  # Get required window width
        height = self.root.winfo_reqheight()  # Get required window height
        self.root.geometry(f"{width}x{height}")  # Resize window to fit content

if __name__ == "__main__":
    root = tk.Tk()
    app = CellMigrationAnalyzerApp(root)
    app.adjust_window_size() # adjust window size to fit content
    root.mainloop()
