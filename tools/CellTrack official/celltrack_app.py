import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.spatial.distance import euclidean
from scipy.stats import ttest_ind, mannwhitneyu
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import io
import base64

# Page configuration
st.set_page_config(
    page_title="CellTrack Official - Enhanced",
    page_icon="🔬",
    layout="wide"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    .stSlider > div > div > div > div {
        background-color: #1f77b4;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown('<h1 class="main-header">CellTrack Official - Enhanced</h1>', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; font-size: 1.2rem; color: #666;">Advanced Cell Migration Analysis with Comprehensive Statistical Analysis</p>', unsafe_allow_html=True)
st.markdown("---")

# Initialize session state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'all_results' not in st.session_state:
    st.session_state.all_results = None
if 'pixel_size' not in st.session_state:
    st.session_state.pixel_size = 1.0
if 'current_sheet' not in st.session_state:
    st.session_state.current_sheet = None
if 'comparison_results' not in st.session_state:
    st.session_state.comparison_results = None

# Enhanced analysis functions
def analyze_cell_track_enhanced(df, pixel_size):
    """Enhanced cell track analysis with comprehensive metrics"""
    results = {}
    df = df.sort_values(by='slice number').copy()

    # Basic calculations
    df['dx'] = df['X coordinate'].diff().fillna(0) * pixel_size
    df['dy'] = df['Y coordinate'].diff().fillna(0) * pixel_size
    df['distance'] = np.sqrt(df['dx']**2 + df['dy']**2).fillna(0)
    df['time_diff'] = df['time (min)'].diff().bfill()
    df['instantaneous_speed'] = df.apply(lambda row: row['distance'] / row['time_diff'] if row['time_diff'] != 0 else 0, axis=1)
    df['direction'] = np.arctan2(df['dy'], df['dx']).fillna(0)
    df['turning_angle'] = df['direction'].diff().fillna(0)

    # Basic metrics
    results['total_distance'] = df['distance'].sum()
    results['net_displacement'] = euclidean(
        (df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size),
        (df['X coordinate'].iloc[-1] * pixel_size, df['Y coordinate'].iloc[-1] * pixel_size)
    )
    results['average_speed'] = df['instantaneous_speed'].mean()
    results['max_speed'] = df['instantaneous_speed'].max()
    results['min_speed'] = df['instantaneous_speed'].min()
    results['speed_std'] = df['instantaneous_speed'].std()

    # Directionality analysis
    results['directionality'] = results['net_displacement'] / results['total_distance'] if results['total_distance'] != 0 else np.nan
    results['path_tortuosity'] = results['total_distance'] / results['net_displacement'] if results['net_displacement'] != 0 else np.nan

    # Persistence analysis
    results['persistence_time'] = (df['turning_angle'].abs() < np.pi/4).sum() * df['time_diff'].mean()
    results['straightness_index'] = results['net_displacement'] / results['total_distance'] if results['total_distance'] != 0 else np.nan

    # Time-based directionality
    directionality_time = []
    for i in range(1, len(df)):
        net_disp = euclidean(
            (df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size),
            (df['X coordinate'].iloc[i] * pixel_size, df['Y coordinate'].iloc[i] * pixel_size)
        )
        total_dist = df['distance'].iloc[:i+1].sum()
        directionality_time.append(net_disp / total_dist if total_dist != 0 else np.nan)
    results['directionality_over_time'] = directionality_time

    # Distance-based directionality
    distance_bins = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    distance_bins = [bin_value * pixel_size for bin_value in distance_bins]
    directionality_distance = []
    total_distance_traveled = 0
    bin_index = 0
    for i in range(1, len(df)):
        total_distance_traveled += df['distance'].iloc[i]
        while bin_index < len(distance_bins) and total_distance_traveled >= distance_bins[bin_index]:
            net_disp = euclidean(
                (df['X coordinate'].iloc[0] * pixel_size, df['Y coordinate'].iloc[0] * pixel_size),
                (df['X coordinate'].iloc[i] * pixel_size, df['Y coordinate'].iloc[i] * pixel_size)
            )
            directionality_distance.append(net_disp / total_distance_traveled if total_distance_traveled != 0 else np.nan)
            bin_index += 1
            if bin_index >= len(distance_bins):
                break
    results['directionality_over_distance'] = directionality_distance

    # Mean Squared Displacement (MSD)
    msd = []
    num_points = len(df)
    for tau in range(1, min(num_points, 20)):  # Limit to first 20 time points
        displacements = [
            (df['X coordinate'].iloc[i+tau] - df['X coordinate'].iloc[i])**2 + 
            (df['Y coordinate'].iloc[i+tau] - df['Y coordinate'].iloc[i])**2 
            for i in range(num_points - tau)
        ]
        msd.append(np.mean(displacements) * pixel_size**2)
    results['mean_squared_displacement'] = msd

    # Additional metrics
    results['track_duration'] = df['time (min)'].max() - df['time (min)'].min()
    results['num_points'] = len(df)
    results['turning_angle_std'] = df['turning_angle'].std()
    results['turning_angle_mean'] = df['turning_angle'].mean()

    # Store detailed data
    results['instantaneous_speed'] = df['instantaneous_speed'].tolist()
    results['direction'] = df['direction'].tolist()
    results['turning_angles'] = df['turning_angle'].tolist()
    results['data'] = df

    return results

def process_excel_file_enhanced(file_path, pixel_size):
    """Process Excel file with enhanced analysis"""
    try:
        xl = pd.ExcelFile(file_path)
        all_results = {}

        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name)
            
            # Validate required columns
            required_columns = ['track number', 'slice number', 'X coordinate', 'Y coordinate', 'time (min)']
            if not all(col in df.columns for col in required_columns):
                st.error(f"Sheet '{sheet_name}' is missing required columns: {required_columns}")
                continue

            track_ids = df['track number'].unique()
            sheet_results = {}

            for track_id in track_ids:
                track_df = df[df['track number'] == track_id].copy()
                if len(track_df) > 1:  # Only analyze tracks with multiple points
                    track_results = analyze_cell_track_enhanced(track_df, pixel_size)
                    sheet_results[track_id] = track_results

            all_results[sheet_name] = sheet_results

        return all_results
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        return None

def compare_conditions_enhanced(all_results):
    """Enhanced condition comparison with statistical tests"""
    comparison_results = {}
    
    # Extract metrics for comparison
    metrics_to_compare = [
        'total_distance', 'net_displacement', 'average_speed', 'directionality',
        'persistence_time', 'path_tortuosity', 'straightness_index'
    ]
    
    for metric in metrics_to_compare:
        condition_data = {}
        
        for sheet_name, sheet_results in all_results.items():
            values = []
            for track_id, track_results in sheet_results.items():
                if metric in track_results and not np.isnan(track_results[metric]):
                    values.append(track_results[metric])
            
            if values:
                condition_data[sheet_name] = values
        
        # Perform statistical tests if we have at least 2 conditions
        if len(condition_data) >= 2:
            conditions = list(condition_data.keys())
            comparison_results[metric] = {
                'conditions': conditions,
                'data': condition_data,
                'means': {cond: np.mean(data) for cond, data in condition_data.items()},
                'stds': {cond: np.std(data) for cond, data in condition_data.items()},
                'counts': {cond: len(data) for cond, data in condition_data.items()}
            }
            
            # Perform t-test for each pair
            if len(conditions) == 2:
                try:
                    t_stat, p_value = ttest_ind(condition_data[conditions[0]], condition_data[conditions[1]])
                    comparison_results[metric]['t_test'] = {
                        't_statistic': t_stat,
                        'p_value': p_value,
                        'significant': p_value < 0.05
                    }
                except:
                    comparison_results[metric]['t_test'] = None
            
            # Perform Mann-Whitney U test for each pair
            if len(conditions) == 2:
                try:
                    u_stat, p_value = mannwhitneyu(condition_data[conditions[0]], condition_data[conditions[1]], alternative='two-sided')
                    comparison_results[metric]['mann_whitney'] = {
                        'u_statistic': u_stat,
                        'p_value': p_value,
                        'significant': p_value < 0.05
                    }
                except:
                    comparison_results[metric]['mann_whitney'] = None
    
    return comparison_results

def create_enhanced_plots(results, sheet_name):
    """Create enhanced visualization plots"""
    plots = {}
    
    # Extract data for plotting
    speeds = []
    directions = []
    turning_angles = []
    directionality_time = []
    
    for track_id, track_results in results.items():
        speeds.extend(track_results['instantaneous_speed'])
        directions.extend(track_results['direction'])
        turning_angles.extend(track_results['turning_angles'])
        directionality_time.extend(track_results['directionality_over_time'])
    
    # 1. Speed distribution
    fig_speed, ax_speed = plt.subplots(figsize=(10, 6))
    ax_speed.hist(speeds, bins=30, alpha=0.7, edgecolor='black')
    ax_speed.set_xlabel('Instantaneous Speed (μm/min)')
    ax_speed.set_ylabel('Frequency')
    ax_speed.set_title(f'Speed Distribution - {sheet_name}')
    ax_speed.grid(True, alpha=0.3)
    plots['speed_distribution'] = fig_speed
    
    # 2. Direction distribution
    fig_direction, ax_direction = plt.subplots(figsize=(10, 6))
    ax_direction.hist(directions, bins=36, alpha=0.7, edgecolor='black')
    ax_direction.set_xlabel('Direction (radians)')
    ax_direction.set_ylabel('Frequency')
    ax_direction.set_title(f'Direction Distribution - {sheet_name}')
    ax_direction.grid(True, alpha=0.3)
    plots['direction_distribution'] = fig_direction
    
    # 3. Turning angle distribution
    fig_turning, ax_turning = plt.subplots(figsize=(10, 6))
    ax_turning.hist(turning_angles, bins=30, alpha=0.7, edgecolor='black')
    ax_turning.set_xlabel('Turning Angle (radians)')
    ax_turning.set_ylabel('Frequency')
    ax_turning.set_title(f'Turning Angle Distribution - {sheet_name}')
    ax_turning.grid(True, alpha=0.3)
    plots['turning_angle_distribution'] = fig_turning
    
    # 4. Directionality over time
    fig_dir_time, ax_dir_time = plt.subplots(figsize=(10, 6))
    directionality_time_clean = [x for x in directionality_time if not np.isnan(x)]
    if directionality_time_clean:
        ax_dir_time.hist(directionality_time_clean, bins=30, alpha=0.7, edgecolor='black')
        ax_dir_time.set_xlabel('Directionality Ratio')
        ax_dir_time.set_ylabel('Frequency')
        ax_dir_time.set_title(f'Directionality Distribution - {sheet_name}')
        ax_dir_time.grid(True, alpha=0.3)
        plots['directionality_distribution'] = fig_dir_time
    
    return plots

def create_comparison_plots(comparison_results):
    """Create comparison plots between conditions"""
    plots = {}
    
    for metric, data in comparison_results.items():
        if 'data' in data:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Create box plot
            box_data = [data['data'][cond] for cond in data['conditions']]
            bp = ax.boxplot(box_data, labels=data['conditions'], patch_artist=True)
            
            # Color the boxes
            colors = ['lightblue', 'lightgreen', 'lightcoral', 'lightyellow']
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)
            
            ax.set_ylabel(metric.replace('_', ' ').title())
            ax.set_title(f'{metric.replace("_", " ").title()} Comparison')
            ax.grid(True, alpha=0.3)
            
            # Add statistical significance markers
            if 't_test' in data and data['t_test']:
                p_value = data['t_test']['p_value']
                significance = '*' if p_value < 0.05 else 'ns'
                ax.text(0.5, 0.95, f'p = {p_value:.3f} {significance}', 
                       transform=ax.transAxes, ha='center', va='top',
                       bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
            
            plots[f'{metric}_comparison'] = fig
    
    return plots

# Main interface
col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("📁 File Upload")
    uploaded_file = st.file_uploader(
        "Choose an Excel file with cell tracking data",
        type=['xlsx', 'xls'],
        help="Upload Excel file with columns: track number, slice number, X coordinate, Y coordinate, time (min)"
    )
    
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success(f"✅ Uploaded: {uploaded_file.name}")

with col2:
    st.subheader("⚙️ Analysis Settings")
    pixel_size = st.number_input(
        "Pixel Size (μm)", 
        min_value=0.1, 
        max_value=10.0, 
        value=1.0, 
        step=0.1,
        help="Physical size of one pixel in micrometers"
    )
    st.session_state.pixel_size = pixel_size

# Analysis section
if st.session_state.uploaded_file is not None:
    st.markdown("---")
    
    # Create tabs for different views
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Data Overview", "🔬 Track Analysis", "📈 Visualizations", "📊 Statistical Comparison", "💾 Export Results"])
    
    with tab1:
        st.subheader("Data Overview")
        
        if st.button("🔍 Analyze Data"):
            with st.spinner("Processing cell tracking data..."):
                # Save uploaded file temporarily
                with open("temp_file.xlsx", "wb") as f:
                    f.write(st.session_state.uploaded_file.getbuffer())
                
                # Process the file
                all_results = process_excel_file_enhanced("temp_file.xlsx", pixel_size)
                
                if all_results:
                    st.session_state.all_results = all_results
                    st.success("✅ Analysis completed!")
                    
                    # Display summary
                    total_tracks = sum(len(sheet_results) for sheet_results in all_results.values())
                    total_sheets = len(all_results)
                    
                    col_sum1, col_sum2, col_sum3 = st.columns(3)
                    with col_sum1:
                        st.metric("Total Sheets", total_sheets)
                    with col_sum2:
                        st.metric("Total Tracks", total_tracks)
                    with col_sum3:
                        st.metric("Pixel Size", f"{pixel_size} μm")
                    
                    # Sheet selection
                    sheet_names = list(all_results.keys())
                    selected_sheet = st.selectbox("Select Sheet for Detailed Analysis", sheet_names)
                    st.session_state.current_sheet = selected_sheet
                    
                    if selected_sheet:
                        sheet_results = all_results[selected_sheet]
                        st.write(f"**Sheet: {selected_sheet}**")
                        st.write(f"- Number of tracks: {len(sheet_results)}")
                        
                        # Display track summary
                        track_summary = []
                        for track_id, track_results in sheet_results.items():
                            track_summary.append({
                                'Track ID': track_id,
                                'Total Distance (μm)': f"{track_results['total_distance']:.2f}",
                                'Net Displacement (μm)': f"{track_results['net_displacement']:.2f}",
                                'Average Speed (μm/min)': f"{track_results['average_speed']:.2f}",
                                'Directionality': f"{track_results['directionality']:.3f}",
                                'Duration (min)': f"{track_results['track_duration']:.1f}"
                            })
                        
                        track_df = pd.DataFrame(track_summary)
                        st.dataframe(track_df, use_container_width=True)
    
    with tab2:
        st.subheader("Individual Track Analysis")
        
        if st.session_state.all_results and st.session_state.current_sheet:
            sheet_results = st.session_state.all_results[st.session_state.current_sheet]
            track_ids = list(sheet_results.keys())
            
            selected_track = st.selectbox("Select Track for Detailed Analysis", track_ids)
            
            if selected_track:
                track_results = sheet_results[selected_track]
                track_data = track_results['data']
                
                col_track1, col_track2 = st.columns(2)
                
                with col_track1:
                    st.write("**Track Metrics:**")
                    st.metric("Total Distance", f"{track_results['total_distance']:.2f} μm")
                    st.metric("Net Displacement", f"{track_results['net_displacement']:.2f} μm")
                    st.metric("Average Speed", f"{track_results['average_speed']:.2f} μm/min")
                    st.metric("Directionality", f"{track_results['directionality']:.3f}")
                    st.metric("Persistence Time", f"{track_results['persistence_time']:.2f} min")
                    st.metric("Path Tortuosity", f"{track_results['path_tortuosity']:.3f}")
                
                with col_track2:
                    st.write("**Track Statistics:**")
                    st.metric("Max Speed", f"{track_results['max_speed']:.2f} μm/min")
                    st.metric("Min Speed", f"{track_results['min_speed']:.2f} μm/min")
                    st.metric("Speed Std", f"{track_results['speed_std']:.2f} μm/min")
                    st.metric("Track Duration", f"{track_results['track_duration']:.1f} min")
                    st.metric("Number of Points", track_results['num_points'])
                    st.metric("Turning Angle Std", f"{track_results['turning_angle_std']:.3f}")
                
                # Track trajectory plot
                fig_traj, ax_traj = plt.subplots(figsize=(10, 8))
                ax_traj.plot(track_data['X coordinate'] * pixel_size, track_data['Y coordinate'] * pixel_size, 'b-o', linewidth=2, markersize=4)
                ax_traj.plot(track_data['X coordinate'].iloc[0] * pixel_size, track_data['Y coordinate'].iloc[0] * pixel_size, 'go', markersize=8, label='Start')
                ax_traj.plot(track_data['X coordinate'].iloc[-1] * pixel_size, track_data['Y coordinate'].iloc[-1] * pixel_size, 'ro', markersize=8, label='End')
                ax_traj.set_xlabel('X Position (μm)')
                ax_traj.set_ylabel('Y Position (μm)')
                ax_traj.set_title(f'Track {selected_track} Trajectory')
                ax_traj.legend()
                ax_traj.grid(True, alpha=0.3)
                st.pyplot(fig_traj)
    
    with tab3:
        st.subheader("Visualization Gallery")
        
        if st.session_state.all_results and st.session_state.current_sheet:
            sheet_results = st.session_state.all_results[st.session_state.current_sheet]
            
            if st.button("📊 Generate Visualizations"):
                with st.spinner("Creating visualizations..."):
                    plots = create_enhanced_plots(sheet_results, st.session_state.current_sheet)
                    
                    # Display plots
                    for plot_name, fig in plots.items():
                        st.write(f"**{plot_name.replace('_', ' ').title()}**")
                        st.pyplot(fig)
                        st.markdown("---")
    
    with tab4:
        st.subheader("Statistical Comparison")
        
        if st.session_state.all_results:
            if st.button("📊 Compare Conditions"):
                with st.spinner("Performing statistical analysis..."):
                    comparison_results = compare_conditions_enhanced(st.session_state.all_results)
                    st.session_state.comparison_results = comparison_results
                    
                    if comparison_results:
                        st.success("✅ Statistical comparison completed!")
                        
                        # Display comparison results
                        for metric, data in comparison_results.items():
                            st.write(f"**{metric.replace('_', ' ').title()}**")
                            
                            col_comp1, col_comp2 = st.columns(2)
                            
                            with col_comp1:
                                # Summary statistics
                                summary_data = []
                                for cond in data['conditions']:
                                    summary_data.append({
                                        'Condition': cond,
                                        'Mean': f"{data['means'][cond]:.3f}",
                                        'Std': f"{data['stds'][cond]:.3f}",
                                        'Count': data['counts'][cond]
                                    })
                                summary_df = pd.DataFrame(summary_data)
                                st.dataframe(summary_df, use_container_width=True)
                            
                            with col_comp2:
                                # Statistical tests
                                if 't_test' in data and data['t_test']:
                                    st.write("**T-Test Results:**")
                                    st.write(f"t-statistic: {data['t_test']['t_statistic']:.3f}")
                                    st.write(f"p-value: {data['t_test']['p_value']:.3f}")
                                    st.write(f"Significant: {'Yes' if data['t_test']['significant'] else 'No'}")
                                
                                if 'mann_whitney' in data and data['mann_whitney']:
                                    st.write("**Mann-Whitney U Test:**")
                                    st.write(f"U-statistic: {data['mann_whitney']['u_statistic']:.3f}")
                                    st.write(f"p-value: {data['mann_whitney']['p_value']:.3f}")
                                    st.write(f"Significant: {'Yes' if data['mann_whitney']['significant'] else 'No'}")
                            
                            st.markdown("---")
                        
                        # Create comparison plots
                        if st.button("📈 Generate Comparison Plots"):
                            comparison_plots = create_comparison_plots(comparison_results)
                            for plot_name, fig in comparison_plots.items():
                                st.write(f"**{plot_name.replace('_', ' ').title()}**")
                                st.pyplot(fig)
                                st.markdown("---")
                    else:
                        st.warning("⚠️ Need at least 2 conditions for statistical comparison")
    
    with tab5:
        st.subheader("Export Results")
        
        if st.session_state.all_results:
            col_exp1, col_exp2 = st.columns(2)
            
            with col_exp1:
                if st.button("📄 Export Summary Report"):
                    # Create summary report
                    report_data = []
                    for sheet_name, sheet_results in st.session_state.all_results.items():
                        for track_id, track_results in sheet_results.items():
                            report_data.append({
                                'Sheet': sheet_name,
                                'Track_ID': track_id,
                                'Total_Distance_μm': track_results['total_distance'],
                                'Net_Displacement_μm': track_results['net_displacement'],
                                'Average_Speed_μm_min': track_results['average_speed'],
                                'Directionality': track_results['directionality'],
                                'Persistence_Time_min': track_results['persistence_time'],
                                'Path_Tortuosity': track_results['path_tortuosity'],
                                'Track_Duration_min': track_results['track_duration'],
                                'Num_Points': track_results['num_points']
                            })
                    
                    report_df = pd.DataFrame(report_data)
                    
                    # Convert to Excel
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        report_df.to_excel(writer, sheet_name='Summary', index=False)
                        
                        # Add color coding for different sheets
                        workbook = writer.book
                        worksheet = writer.sheets['Summary']
                        
                        # Color coding based on sheet
                        sheet_colors = ['FFE6E6', 'E6F3FF', 'E6FFE6', 'FFF2E6']
                        for i, sheet_name in enumerate(report_df['Sheet'].unique()):
                            color = sheet_colors[i % len(sheet_colors)]
                            for row in range(2, len(report_df) + 2):
                                if worksheet.cell(row=row, column=1).value == sheet_name:
                                    for col in range(1, len(report_df.columns) + 1):
                                        cell = worksheet.cell(row=row, column=col)
                                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    output.seek(0)
                    st.download_button(
                        label="📥 Download Summary Report",
                        data=output.getvalue(),
                        file_name="cell_tracking_summary.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            with col_exp2:
                if st.button("📊 Export Comparison Results"):
                    if st.session_state.comparison_results:
                        # Create comparison report
                        comp_data = []
                        for metric, data in st.session_state.comparison_results.items():
                            for cond in data['conditions']:
                                comp_data.append({
                                    'Metric': metric,
                                    'Condition': cond,
                                    'Mean': data['means'][cond],
                                    'Std': data['stds'][cond],
                                    'Count': data['counts'][cond]
                                })
                        
                        comp_df = pd.DataFrame(comp_data)
                        
                        # Convert to Excel
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            comp_df.to_excel(writer, sheet_name='Comparison', index=False)
                        
                        output.seek(0)
                        st.download_button(
                            label="📥 Download Comparison Results",
                            data=output.getvalue(),
                            file_name="cell_tracking_comparison.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; font-size: 0.9rem;">
    <p>🔬 <strong>CellTrack Official - Enhanced Version</strong></p>
    <p>Advanced features: Comprehensive statistical analysis, multiple visualization options, detailed metrics, and export capabilities</p>
    <p>For the full desktop experience with all features, download the original Tkinter application</p>
    <p>📧 Contact: <a href="mailto:jalal.rahbani@hotmail.com">jalal.rahbani@hotmail.com</a> | 
    🌐 <a href="https://jalalrahbani.github.io/-jalal-website/" target="_blank">Personal Website</a></p>
</div>
""", unsafe_allow_html=True) 